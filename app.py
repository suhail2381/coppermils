"""
==========================================================================================
 ELECTRIC WIRE & CABLE — SALES + PRODUCTION + INVENTORY PORTAL
==========================================================================================
A single-file Streamlit application for a wire/cable manufacturing business covering:
  • Public storefront (customers browse products & place orders — no login required)
  • Customer records (auto-created from orders)
  • Sales / Order management
  • Raw material inventory (Copper: Pure / Loose Gauge / China · PVC & XLPE grades · Aluminium)
  • Production entry (labour, supervisor, production manager, output & rejected/broken qty)
  • Scrap / condemned material tracking (copper & PVC separately)
  • Product catalog with editable pricing
  • Staff records (labour / supervisor / production manager)
  • Consolidated stock & sales reports with charts and Excel export
==========================================================================================
"""

import streamlit as st
import sqlite3
import pandas as pd
import plotly.express as px
import io
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------------------------------
# 0. APP CONFIG & BUSINESS SETTINGS
# ------------------------------------------------------------------------------------
BUSINESS_NAME = "Apna Wire & Cable Industries"
CURRENCY = "PKR"
DB_FILE = "wire_cable_erp.db"

st.set_page_config(page_title=f"{BUSINESS_NAME} — Portal", layout="wide", page_icon="⚡")

# ------------------------------------------------------------------------------------
# 1. GLOBAL STYLING — industrial copper theme
# ------------------------------------------------------------------------------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
    :root {
        --copper: #b45309;
        --copper-dark: #7c2d12;
        --navy: #0f172a;
        --navy-2: #1e293b;
        --surface: #ffffff;
        --surface-alt: #f6f4f1;
        --border-soft: #e7e2da;
        --text-primary: #1c1917;
        --text-muted: #78716c;
        --radius-lg: 18px; --radius-md: 12px; --radius-sm: 8px;
        --shadow-soft: 0 4px 16px rgba(28,25,23,0.06);
        --shadow-lift: 0 12px 28px rgba(28,25,23,0.10);
    }
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .stApp {
        background: radial-gradient(circle at 0% 0%, rgba(180,83,9,0.05) 0%, transparent 45%),
                    radial-gradient(circle at 100% 0%, rgba(15,23,42,0.04) 0%, transparent 45%),
                    var(--surface-alt);
    }
    .block-container { padding-top: 1.4rem; max-width: 1400px; }
    h1, h2, h3 { color: var(--text-primary); letter-spacing: -0.01em; }
    hr { border-color: var(--border-soft) !important; }

    div.stButton > button, div.stDownloadButton > button, div.stFormSubmitButton > button {
        background: linear-gradient(135deg, var(--copper) 0%, var(--copper-dark) 100%) !important;
        color: white !important; border: none !important; padding: 0.65rem 1.4rem !important;
        font-weight: 600 !important; border-radius: var(--radius-sm) !important;
        box-shadow: 0 6px 16px rgba(180,83,9,0.25) !important; transition: all .2s ease !important;
    }
    div.stButton > button:hover, div.stDownloadButton > button:hover, div.stFormSubmitButton > button:hover {
        transform: translateY(-2px); box-shadow: 0 10px 22px rgba(180,83,9,0.32) !important;
    }

    div[data-testid="stForm"] {
        background: var(--surface) !important; border-radius: var(--radius-lg) !important;
        padding: 26px !important; box-shadow: var(--shadow-soft) !important;
        border: 1px solid var(--border-soft) !important;
    }

    .kpi-card { border-radius: var(--radius-md); padding: 20px; color: white; box-shadow: var(--shadow-soft);
                transition: transform .2s ease; }
    .kpi-card:hover { transform: translateY(-3px); }
    .kpi-title { font-size: 11.5px; font-weight: 700; text-transform: uppercase; letter-spacing: .8px; opacity: .85; }
    .kpi-value { font-size: 24px; font-weight: 800; margin-top: 6px; }
    .kpi-copper { background: linear-gradient(135deg, #7c2d12, #b45309); }
    .kpi-navy   { background: linear-gradient(135deg, #0f172a, #1e3a5f); }
    .kpi-green  { background: linear-gradient(135deg, #064e3b, #10b981); }
    .kpi-red    { background: linear-gradient(135deg, #7f1d1d, #ef4444); }
    .kpi-slate  { background: linear-gradient(135deg, #334155, #64748b); }

    .product-card { background: var(--surface); border: 1px solid var(--border-soft); border-radius: var(--radius-md);
                     padding: 16px; box-shadow: var(--shadow-soft); margin-bottom: 14px; transition: transform .15s ease; }
    .product-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-lift); }
    .product-tag { display:inline-block; font-size:11px; font-weight:700; padding:3px 10px; border-radius:20px;
                    background: rgba(180,83,9,0.12); color: var(--copper-dark); margin-right:6px; }
    .stock-ok   { color:#15803d; font-weight:700; font-size:12.5px; }
    .stock-low  { color:#b45309; font-weight:700; font-size:12.5px; }
    .stock-out  { color:#b91c1c; font-weight:700; font-size:12.5px; }

    .stTabs [data-baseweb="tab"] { border-radius: var(--radius-sm); font-weight:600; }
    .stTabs [aria-selected="true"] { background: linear-gradient(135deg, var(--copper), var(--copper-dark)) !important; color:white !important; }

    section[data-testid="stSidebar"] { background: linear-gradient(180deg, var(--navy) 0%, var(--navy-2) 100%); }
    section[data-testid="stSidebar"] * { color: #e7e5e4; }
    section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.12) !important; }
</style>
""", unsafe_allow_html=True)

# ------------------------------------------------------------------------------------
# 2. DATABASE LAYER (THREAD-SAFE SETUP)
# ------------------------------------------------------------------------------------
def get_conn():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL;")
    return conn

def init_db():
    with get_conn() as conn:
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS products (
            product_id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT, material TEXT, spec_name TEXT, core_config TEXT,
            flexibility TEXT, material_quality TEXT, insulation_quality TEXT,
            application TEXT, unit TEXT, price REAL DEFAULT 0, stock_qty REAL DEFAULT 0,
            reorder_level REAL DEFAULT 10, status TEXT DEFAULT 'Active')""")

        c.execute("""CREATE TABLE IF NOT EXISTS price_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER, old_price REAL,
            new_price REAL, changed_on TEXT, changed_by TEXT)""")

        c.execute("""CREATE TABLE IF NOT EXISTS customers (
            customer_id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, phone TEXT UNIQUE,
            email TEXT, address TEXT, created_on TEXT)""")

        c.execute("""CREATE TABLE IF NOT EXISTS orders (
            order_id INTEGER PRIMARY KEY AUTOINCREMENT, customer_id INTEGER, order_date TEXT,
            status TEXT DEFAULT 'Pending', total_amount REAL DEFAULT 0, notes TEXT)""")

        c.execute("""CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, order_id INTEGER, product_id INTEGER,
            qty REAL, unit_price REAL, subtotal REAL)""")

        c.execute("""CREATE TABLE IF NOT EXISTS raw_materials (
            material_id INTEGER PRIMARY KEY AUTOINCREMENT, material_type TEXT, category TEXT,
            unit TEXT DEFAULT 'kg', stock_qty REAL DEFAULT 0, reorder_level REAL DEFAULT 50,
            unit_cost REAL DEFAULT 0)""")

        c.execute("""CREATE TABLE IF NOT EXISTS raw_material_txn (
            id INTEGER PRIMARY KEY AUTOINCREMENT, material_id INTEGER, txn_type TEXT,
            qty REAL, txn_date TEXT, reference TEXT, notes TEXT)""")

        c.execute("""CREATE TABLE IF NOT EXISTS staff (
            staff_id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, role TEXT, contact TEXT,
            join_date TEXT, status TEXT DEFAULT 'Active')""")

        c.execute("""CREATE TABLE IF NOT EXISTS production (
            production_id INTEGER PRIMARY KEY AUTOINCREMENT, production_date TEXT, product_id INTEGER,
            qty_produced REAL, qty_rejected REAL DEFAULT 0, supervisor_id INTEGER,
            manager_id INTEGER, labour_count INTEGER DEFAULT 0, notes TEXT)""")

        c.execute("""CREATE TABLE IF NOT EXISTS production_material_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT, production_id INTEGER, material_id INTEGER, qty_used REAL)""")

        c.execute("""CREATE TABLE IF NOT EXISTS scrap_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, log_date TEXT, material_category TEXT,
            sub_category TEXT, qty REAL, unit TEXT DEFAULT 'kg', reason TEXT, notes TEXT)""")

def seed_reference_data():
    with get_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM products")
        if c.fetchone()[0] == 0:
            products = [
                ("Building Wire", "Aluminium", "7/36 Twin Core (2-Core)", "Double Core / Flat Sheathed", "Stiff / Rigid", "EC Grade Aluminium (99.5%+)", "Weatherproof PVC (UV-stabilized)", "Low-load outdoor lights, small shops", "90m Coil", 3200, 40, 10),
                ("Building Wire", "Aluminium", "7/44 Twin Core (2-Core)", "Double Core / Flat Sheathed", "Stiff / Rigid", "EC Grade Aluminium (99.5%+)", "Weatherproof PVC (UV-stabilized)", "Sub-mains, single floor supply", "90m Coil", 3800, 35, 10),
                ("Building Wire", "Aluminium", "7/52 Twin Core (2-Core)", "Double / Round Sheathed", "Very Rigid", "EC Grade Aluminium (99.5%+)", "Double Layer Heavy PVC", "Standard single-phase main line", "90m Coil", 4600, 30, 8),
                ("Building Wire", "Aluminium", "7/64 Twin Core (2-Core)", "Double / Round Sheathed", "Very Rigid", "EC Grade Aluminium (99.5%+)", "Double Layer Heavy PVC", "Heavy single-phase / 3-phase meter lines", "90m Coil", 5400, 25, 8),
                ("Building Wire", "Aluminium", "19/52 Single or 2-Core", "Heavy Service Mains", "Extremely Rigid", "EC Grade Aluminium (99.5%+)", "XLPE / Heavy PVC Casing", "3-Phase commercial connections", "90m Coil", 7200, 20, 5),
                ("Building Wire", "Aluminium", "19/64 Single or 2-Core", "Heavy Service Mains", "Extremely Rigid", "EC Grade Aluminium (99.5%+)", "XLPE / Heavy PVC Casing", "Heavy commercial / tube wells", "90m Coil", 8600, 15, 5),
                ("Building Wire", "Aluminium", "19/83 Single Core", "Main Transmission Feeder", "Extremely Rigid", "EC Grade Aluminium (99.5%+)", "Weatherproof Heavy XLPE", "Industrial / long-distance rural feeds", "90m Coil", 10500, 10, 5),
                ("Building Wire", "Copper", "03/29 Size", "Single Core", "Semi-Rigid", "99.9% Pure Electrolytic Copper", "FR-PVC (70°C rated)", "Ceiling fans, LED lights", "90m Coil", 2800, 60, 15),
                ("Building Wire", "Copper", "07/29 Size", "Single Core", "Semi-Rigid", "99.9% Pure Electrolytic Copper", "FR-PVC (70°C rated)", "Wall plugs, refrigerators, inverter ACs", "90m Coil", 4200, 55, 15),
                ("Building Wire", "Copper", "07/36 Size", "Single/Multi-Core Sheathed", "Medium Flexibility", "99.9% Pure Electrolytic Copper", "HR-PVC (85-90°C)", "2-ton ACs, water pumps, washing machines", "90m Coil", 6100, 45, 12),
                ("Building Wire", "Copper", "07/44 Size", "Single/Multi-Core Sheathed", "Medium Flexibility", "99.9% Pure Electrolytic Copper", "HR-PVC (85-90°C)", "Geysers, meter to sub-DB mains", "90m Coil", 7500, 40, 12),
                ("Building Wire", "Copper", "07/52 Size", "Single/Twin Core", "Stiff / Rigid", "99.9% Pure Electrolytic Copper", "Heavy XLPE / PVC Double Jacket", "Meter to main breaker panel", "90m Coil", 9200, 30, 10),
                ("Building Wire", "Copper", "07/64 Size", "Single/Multi-Core", "Stiff / Rigid", "99.9% Pure Electrolytic Copper", "Heavy XLPE / PVC Double Jacket", "Service entrance, multi-story buildings", "90m Coil", 11800, 20, 8),
                ("Building Wire", "Copper", "23/76 Size", "Flexible Multi-Core Cord", "High Flexibility", "99.9% Stranded Copper (annealed)", "Super Flexible PVC", "Pendant lights, desk electronics", "90m Coil", 5300, 25, 8),
                ("Building Wire", "Copper", "40/76 Size", "Flexible Multi-Core Cord", "High Flexibility", "99.9% Stranded Copper (annealed)", "Flexible PVC", "Light-duty flexible cords", "90m Coil", 6000, 20, 8),
                ("Power Cable", "Copper", "3-Core Armoured Power Cable", "Multi-Core Armoured", "Rigid", "99.9% Pure Electrolytic Copper", "XLPE Insulated / PVC Sheathed", "Industrial & underground power feeds", "Per Meter", 850, 500, 100),
                ("Power Cable", "Copper", "4-Core Armoured Power Cable", "Multi-Core Armoured", "Rigid", "99.9% Pure Electrolytic Copper", "XLPE Insulated / PVC Sheathed", "3-phase industrial power distribution", "Per Meter", 1050, 400, 100),
                ("Networking Cable (Cat6)", "Copper", "Cat6 UTP Indoor (305m Box)", "4-Pair Twisted", "Flexible", "99.9% Bare Copper Conductor", "PVC / LSZH Jacket", "Structured data & LAN networking", "305m Box", 13500, 25, 5),
                ("Networking Cable (Cat6)", "Copper", "Cat6 FTP Outdoor (305m Box)", "4-Pair Twisted, Foil Shielded", "Flexible", "99.9% Bare Copper Conductor", "UV-Resistant PE Jacket", "Outdoor / campus networking runs", "305m Box", 16500, 15, 5),
            ]
            c.executemany("""INSERT INTO products (category, material, spec_name, core_config, flexibility,
                              material_quality, insulation_quality, application, unit, price, stock_qty, reorder_level)
                              VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", products)

        c.execute("SELECT COUNT(*) FROM raw_materials")
        if c.fetchone()[0] == 0:
            materials = [
                ("Copper", "Pure Copper (Electrolytic 99.9%)", "kg", 500, 100, 2200),
                ("Copper", "Loose Gauge Copper", "kg", 300, 80, 1800),
                ("Copper", "China Copper", "kg", 200, 60, 1500),
                ("Aluminium", "EC Grade Aluminium", "kg", 400, 100, 850),
                ("PVC", "FR-PVC Compound (Flame Retardant)", "kg", 350, 80, 320),
                ("PVC", "HR-PVC Compound (Heat Resistant)", "kg", 300, 80, 350),
                ("PVC", "Weatherproof PVC Compound", "kg", 250, 60, 340),
                ("XLPE", "XLPE Compound (Heavy/HV Grade)", "kg", 250, 60, 420),
            ]
            c.executemany("""INSERT INTO raw_materials (material_type, category, unit, stock_qty, reorder_level, unit_cost)
                              VALUES (?,?,?,?,?,?)""", materials)

init_db()
seed_reference_data()

# ------------------------------------------------------------------------------------
# 3. SECURE USER AUTHENTICATION
# ------------------------------------------------------------------------------------
def load_user_db():
    if "users" in st.secrets:
        return {u: dict(v) for u, v in st.secrets["users"].items()}
    return {
        "admin":   {"password": "changeme-admin",   "role": "Admin",              "display_name": "Owner / Admin"},
        "manager": {"password": "changeme-manager",  "role": "Production Manager", "display_name": "Production Manager"},
        "sales":   {"password": "changeme-sales",    "role": "Sales",              "display_name": "Sales Staff"},
    }

USER_DB = load_user_db()

def staff_login():
    if "staff_authenticated" not in st.session_state:
        st.session_state["staff_authenticated"] = False
        st.session_state["staff_role"] = None

    if st.session_state["staff_authenticated"]:
        return True

    _, col, _ = st.columns([1, 1.1, 1])
    with col:
        with st.form("staff_login_form"):
            st.markdown('<div class="product-tag">Staff & Admin Portal</div>', unsafe_allow_html=True)
            st.markdown("### 🔐 Sign in to manage the business")
            u = st.text_input("Username").strip().lower()
            p = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", use_container_width=True)
            if submitted:
                if u in USER_DB and USER_DB[u]["password"] == p:
                    st.session_state["staff_authenticated"] = True
                    st.session_state["staff_role"] = USER_DB[u]["role"]
                    st.session_state["staff_name"] = USER_DB[u]["display_name"]
                    st.rerun()
                else:
                    st.error("Invalid username or password.")
    return False

# ------------------------------------------------------------------------------------
# 4. SHARED DATABASE HELPERS
# ------------------------------------------------------------------------------------
def df_query(sql, params=()):
    with get_conn() as conn:
        return pd.read_sql_query(sql, conn, params=params)

def run_execute(sql, params=()):
    with get_conn() as conn:
        c = conn.cursor()
        c.execute(sql, params)
        conn.commit()
        return c.lastrowid

def stock_badge(qty, reorder):
    if qty <= 0:
        return '<span class="stock-out">● Out of Stock</span>'
    elif qty <= reorder:
        return '<span class="stock-low">● Low Stock</span>'
    return '<span class="stock-ok">● In Stock</span>'

def to_excel_bytes(df, sheet_name="Report"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    header_fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Border(*(Side(style="thin", color="CCCCCC"),) * 4)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = header_font; cell.fill = header_fill; cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, len(str(col)) + 4)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def get_or_create_customer(name, phone, email, address):
    with get_conn() as conn:
        c = conn.cursor()
        c.execute("SELECT customer_id FROM customers WHERE phone = ?", (phone,))
        row = c.fetchone()
        if row:
            cust_id = row[0]
            c.execute("UPDATE customers SET name=?, email=?, address=? WHERE customer_id=?",
                      (name, email, address, cust_id))
        else:
            c.execute("INSERT INTO customers (name, phone, email, address, created_on) VALUES (?,?,?,?,?)",
                      (name, phone, email, address, datetime.now().strftime("%Y-%m-%d %H:%M")))
            cust_id = c.lastrowid
        conn.commit()
        return cust_id

# ======================================================================================
# 5. TOP-LEVEL NAVIGATION
# ======================================================================================
if "cart" not in st.session_state:
    st.session_state["cart"] = {}

mode = st.sidebar.radio("Choose Portal", ["🛍️ Customer Storefront", "🔐 Staff / Admin Portal"])
st.sidebar.markdown("---")

# ======================================================================================
# 6. CUSTOMER STOREFRONT
# ======================================================================================
if mode == "🛍️ Customer Storefront":
    l1, l2 = st.columns([0.1, 0.9])
    with l1:
        st.markdown("<h1 style='margin:0;'>⚡</h1>", unsafe_allow_html=True)
    with l2:
        st.title(BUSINESS_NAME)
        st.caption("Power Cables • Building Wires • Networking Cat6 Cables — direct from the factory")
    st.markdown("---")

    products_df = df_query("SELECT * FROM products WHERE status='Active'")

    with st.sidebar:
        st.markdown("### 🔍 Browse Products")
        cat_filter = st.multiselect("Category", sorted(products_df["category"].unique()))
        mat_filter = st.multiselect("Material", sorted(products_df["material"].unique()))
        st.markdown("---")
        st.markdown("### 🛒 Your Cart")
        cart = st.session_state["cart"]
        if cart:
            cart_total = 0
            for pid, qty in list(cart.items()):
                prow = products_df[products_df["product_id"] == pid]
                if prow.empty:
                    continue
                prow = prow.iloc[0]
                line_total = qty * prow["price"]
                cart_total += line_total
                st.write(f"**{prow['spec_name']}** — {qty} x {prow['unit']} = {CURRENCY} {line_total:,.0f}")
                if st.button("Remove", key=f"rm_{pid}"):
                    del st.session_state["cart"][pid]
                    st.rerun()
            st.markdown(f"**Total: {CURRENCY} {cart_total:,.0f}**")
        else:
            st.caption("Cart is empty. Add products below.")

    filtered = products_df.copy()
    if cat_filter:
        filtered = filtered[filtered["category"].isin(cat_filter)]
    if mat_filter:
        filtered = filtered[filtered["material"].isin(mat_filter)]

    st.subheader(f"📦 Available Products ({len(filtered)})")
    cols = st.columns(3)
    for i, (_, p) in enumerate(filtered.iterrows()):
        with cols[i % 3]:
            st.markdown(f"""
            <div class="product-card">
                <span class="product-tag">{p['category']}</span><span class="product-tag">{p['material']}</span>
                <h4 style="margin:10px 0 4px 0;">{p['spec_name']}</h4>
                <div style="font-size:12.5px; color:#78716c; margin-bottom:8px;">{p['application']}</div>
                <div style="font-size:20px; font-weight:800; color:#7c2d12;">{CURRENCY} {p['price']:,.0f} <span style="font-size:12px; color:#78716c; font-weight:500;">/ {p['unit']}</span></div>
                <div style="margin-top:6px;">{stock_badge(p['stock_qty'], p['reorder_level'])}</div>
            </div>
            """, unsafe_allow_html=True)
            qty = st.number_input("Qty", min_value=1, max_value=max(1, int(p["stock_qty"])) if p["stock_qty"] > 0 else 1,
                                   value=1, key=f"qty_{p['product_id']}", disabled=(p["stock_qty"] <= 0))
            if st.button("➕ Add to Cart", key=f"add_{p['product_id']}", disabled=(p["stock_qty"] <= 0), use_container_width=True):
                st.session_state["cart"][p["product_id"]] = st.session_state["cart"].get(p["product_id"], 0) + qty
                st.success(f"Added {qty} x {p['spec_name']} to cart.")

    st.markdown("---")
    st.subheader("✅ Checkout")
    if not st.session_state["cart"]:
        st.info("Add at least one product to your cart before checking out.")
    else:
        with st.form("checkout_form"):
            c1, c2 = st.columns(2)
            with c1:
                cust_name = st.text_input("Full Name *")
                cust_phone = st.text_input("Phone Number *")
            with c2:
                cust_email = st.text_input("Email (optional)")
                cust_address = st.text_area("Delivery Address *", height=80)
            order_notes = st.text_input("Order Notes (optional)")
            place_order = st.form_submit_button("🚀 Place Order", use_container_width=True)

        if place_order:
            if not cust_name or not cust_phone or not cust_address:
                st.error("Please fill in Name, Phone, and Address to place your order.")
            else:
                cust_id = get_or_create_customer(cust_name, cust_phone, cust_email, cust_address)
                order_total = 0
                order_id = run_execute(
                    "INSERT INTO orders (customer_id, order_date, status, total_amount, notes) VALUES (?,?,?,?,?)",
                    (cust_id, datetime.now().strftime("%Y-%m-%d %H:%M"), "Pending", 0, order_notes))
                for pid, qty in st.session_state["cart"].items():
                    prow = products_df[products_df["product_id"] == pid].iloc[0]
                    subtotal = qty * prow["price"]
                    order_total += subtotal
                    run_execute(
                        "INSERT INTO order_items (order_id, product_id, qty, unit_price, subtotal) VALUES (?,?,?,?,?)",
                        (order_id, pid, qty, prow["price"], subtotal))
                run_execute("UPDATE orders SET total_amount=? WHERE order_id=?", (order_total, order_id))
                st.session_state["cart"] = {}
                st.success(f"🎉 Order #{order_id} placed successfully! Total: {CURRENCY} {order_total:,.0f}. "
                           f"Our team will contact you shortly to confirm delivery.")

# ======================================================================================
# 7. STAFF / ADMIN PORTAL
# ======================================================================================
else:
    if not staff_login():
        st.stop()

    role = st.session_state["staff_role"]
    st.sidebar.markdown(f"👤 **{st.session_state['staff_name']}**")
    st.sidebar.markdown(f"🏷️ Role: `{role}`")
    if st.sidebar.button("🚪 Log Out", use_container_width=True):
        st.session_state["staff_authenticated"] = False
        st.rerun()
    st.sidebar.markdown("---")

    all_pages = ["📊 Dashboard", "📦 Product Catalog & Pricing", "🧵 Raw Materials", "🏭 Production Entry",
                 "🗑️ Scrap / Condemned Material", "🧾 Orders & Sales", "👥 Customers", "🧑‍🏭 Staff", "📈 Reports"]
    role_pages = {
        "Admin": all_pages,
        "Production Manager": ["📊 Dashboard", "🧵 Raw Materials", "🏭 Production Entry",
                                "🗑️ Scrap / Condemned Material", "🧑‍🏭 Staff", "📈 Reports"],
        "Sales": ["📊 Dashboard", "📦 Product Catalog & Pricing", "🧾 Orders & Sales", "👥 Customers", "📈 Reports"],
    }
    pages = role_pages.get(role, all_pages)
    page = st.sidebar.radio("Navigation", pages)

    st.title(f"⚡ {BUSINESS_NAME} — Staff Portal")
    st.markdown("---")

    # ---------------------------- DASHBOARD ----------------------------
    if page == "📊 Dashboard":
        products_df = df_query("SELECT * FROM products")
        rm_df = df_query("SELECT * FROM raw_materials")
        orders_df = df_query("SELECT * FROM orders")
        scrap_df = df_query("SELECT * FROM scrap_log")

        total_stock_value = (products_df["price"] * products_df["stock_qty"]).sum()
        low_stock_count = (products_df["stock_qty"] <= products_df["reorder_level"]).sum()
        pending_orders = (orders_df["status"] == "Pending").sum() if not orders_df.empty else 0
        today_str = date.today().strftime("%Y-%m-%d")
        today_sales = orders_df[orders_df["order_date"].str.startswith(today_str)]["total_amount"].sum() if not orders_df.empty else 0
        total_scrap = scrap_df["qty"].sum() if not scrap_df.empty else 0

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.markdown(f'<div class="kpi-card kpi-copper"><div class="kpi-title">📦 Stock Value</div><div class="kpi-value">{CURRENCY} {total_stock_value:,.0f}</div></div>', unsafe_allow_html=True)
        k2.markdown(f'<div class="kpi-card kpi-red"><div class="kpi-title">⚠️ Low Stock</div><div class="kpi-value">{low_stock_count}</div></div>', unsafe_allow_html=True)
        k3.markdown(f'<div class="kpi-card kpi-navy"><div class="kpi-title">🧾 Pending Orders</div><div class="kpi-value">{pending_orders}</div></div>', unsafe_allow_html=True)
        k4.markdown(f'<div class="kpi-card kpi-green"><div class="kpi-title">💰 Today\'s Sales</div><div class="kpi-value">{CURRENCY} {today_sales:,.0f}</div></div>', unsafe_allow_html=True)
        k5.markdown(f'<div class="kpi-card kpi-slate"><div class="kpi-title">🗑️ Total Scrap</div><div class="kpi-value">{total_scrap:,.0f} kg</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("##### Stock Value by Category")
            if not products_df.empty:
                cat_val = products_df.groupby("category", as_index=False).apply(
                    lambda d: pd.Series({'value': (d["price"] * d["stock_qty"]).sum()})
                )
                fig = px.bar(cat_val, x="category", y="value", color="category", template="plotly_white")
                fig.update_layout(showlegend=False, height=340)
                st.plotly_chart(fig, use_container_width=True)
        with c2:
            st.markdown("##### Raw Material Stock (kg)")
            if not rm_df.empty:
                fig2 = px.pie(rm_df, names="category", values="stock_qty", hole=0.5, template="plotly_white")
                fig2.update_layout(height=340)
                st.plotly_chart(fig2, use_container_width=True)

    # ---------------------------- PRODUCT CATALOG & PRICING ----------------------------
    elif page == "📦 Product Catalog & Pricing":
        st.subheader("📦 Product Catalog")
        products_df = df_query("SELECT * FROM products ORDER BY category, material, spec_name")
        st.dataframe(products_df.drop(columns=["core_config", "flexibility", "material_quality", "insulation_quality"]),
                     use_container_width=True, hide_index=True)

        st.markdown("---")
        st.subheader("💲 Update Price / Stock Levels")
        with st.form("price_update_form"):
            prod_options = {f"{r['spec_name']} ({r['category']})": r["product_id"] for _, r in products_df.iterrows()}
            selected_label = st.selectbox("Select Product", list(prod_options.keys()))
            selected_id = prod_options[selected_label]
            current = products_df[products_df["product_id"] == selected_id].iloc[0]
            new_price = st.number_input(f"New Price (current: {CURRENCY} {current['price']:,.0f})", min_value=0.0, value=float(current["price"]), step=50.0)
            new_reorder = st.number_input("Reorder Level", min_value=0.0, value=float(current["reorder_level"]), step=1.0)
            update_btn = st.form_submit_button("Update", use_container_width=True)

        if update_btn:
            if new_price != current["price"]:
                run_execute("INSERT INTO price_history (product_id, old_price, new_price, changed_on, changed_by) VALUES (?,?,?,?,?)",
                            (int(selected_id), float(current["price"]), float(new_price),
                             datetime.now().strftime("%Y-%m-%d %H:%M"), st.session_state["staff_name"]))
            run_execute("UPDATE products SET price=?, reorder_level=? WHERE product_id=?",
                        (new_price, new_reorder, int(selected_id)))
            st.success(f"Updated {selected_label}.")
            st.rerun()

        with st.expander("📜 Price Change History"):
            hist_df = df_query("""SELECT ph.changed_on, p.spec_name, ph.old_price, ph.new_price, ph.changed_by
                                   FROM price_history ph JOIN products p ON ph.product_id = p.product_id
                                   ORDER BY ph.id DESC""")
            st.dataframe(hist_df, use_container_width=True, hide_index=True)

    # ---------------------------- RAW MATERIALS ----------------------------
    elif page == "🧵 Raw Materials":
        st.subheader("🧵 Raw Material Stock")
        rm_df = df_query("SELECT * FROM raw_materials ORDER BY material_type, category")
        rm_df["Stock Status"] = rm_df.apply(lambda r: "🔴 Reorder Now" if r["stock_qty"] <= r["reorder_level"] else "🟢 OK", axis=1)
        st.dataframe(rm_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.subheader("📥 Log a Raw Material Transaction")
        with st.form("rm_txn_form", clear_on_submit=True):
            mat_options = {f"{r['material_type']} — {r['category']}": r["material_id"] for _, r in rm_df.iterrows()}
            c1, c2, c3 = st.columns(3)
            with c1:
                sel_mat = st.selectbox("Material", list(mat_options.keys()))
                txn_type = st.selectbox("Transaction Type", ["Purchase", "Consumption", "Scrap / Return", "Adjustment"])
            with c2:
                txn_qty = st.number_input("Quantity (kg)", min_value=0.0, step=1.0)
                txn_date_in = st.date_input("Date", value=date.today())
            with c3:
                reference = st.text_input("Reference (invoice/PO no.)")
                notes = st.text_input("Notes")
            log_btn = st.form_submit_button("Log Transaction", use_container_width=True)

        if log_btn:
            if txn_qty <= 0:
                st.error("Quantity must be greater than zero.")
            else:
                mat_id = mat_options[sel_mat]
                run_execute("INSERT INTO raw_material_txn (material_id, txn_type, qty, txn_date, reference, notes) VALUES (?,?,?,?,?,?)",
                            (mat_id, txn_type, txn_qty, txn_date_in.strftime("%Y-%m-%d"), reference, notes))
                delta = txn_qty if txn_type in ("Purchase", "Adjustment") else -txn_qty
                run_execute("UPDATE raw_materials SET stock_qty = stock_qty + ? WHERE material_id=?", (delta, mat_id))
                st.success(f"Logged {txn_type} of {txn_qty} kg for {sel_mat}.")
                st.rerun()

    # ---------------------------- PRODUCTION ENTRY ----------------------------
    elif page == "🏭 Production Entry":
        st.subheader("🏭 Log a Production Run")
        products_df = df_query("SELECT * FROM products WHERE status='Active'")
        rm_df = df_query("SELECT * FROM raw_materials")
        staff_df = df_query("SELECT * FROM staff WHERE status='Active'")

        supervisors = staff_df[staff_df["role"] == "Supervisor"]
        managers = staff_df[staff_df["role"] == "Production Manager"]

        if "prod_material_rows" not in st.session_state:
            st.session_state["prod_material_rows"] = 1

        prod_options = {f"{r['spec_name']} ({r['category']})": r["product_id"] for _, r in products_df.iterrows()}
        mat_options = {f"{r['material_type']} — {r['category']}": r["material_id"] for _, r in rm_df.iterrows()}
        sup_options = {f"{r['name']}": r["staff_id"] for _, r in supervisors.iterrows()}
        mgr_options = {f"{r['name']}": r["staff_id"] for _, r in managers.iterrows()}

        c1, c2 = st.columns(2)
        with c1:
            prod_date_in = st.date_input("Production Date", value=date.today(), key="pd_date")
            sel_product_label = st.selectbox("Product Produced", list(prod_options.keys()) if prod_options else ["None"], key="pd_prod")
            qty_produced_in = st.number_input("Quantity Produced", min_value=0.0, step=1.0, key="pd_qty")
            qty_rejected_in = st.number_input("Quantity Rejected / Broken", min_value=0.0, step=1.0, key="pd_rej")
        with c2:
            sel_supervisor = st.selectbox("Supervisor", list(sup_options.keys()) if sup_options else ["None"], key="pd_sup")
            sel_manager = st.selectbox("Production Manager", list(mgr_options.keys()) if mgr_options else ["None"], key="pd_mgr")
            labour_count_in = st.number_input("Labour Count", min_value=0, step=1, key="pd_lab")
            prod_notes_in = st.text_input("Notes", key="pd_notes")

        st.markdown("##### Raw Materials Consumed")
        n_rows = st.session_state["prod_material_rows"]
        material_rows = []
        for i in range(n_rows):
            rc1, rc2 = st.columns([2, 1])
            with rc1:
                m_label = st.selectbox(f"Material #{i+1}", list(mat_options.keys()) if mat_options else ["None"], key=f"mat_sel_{i}")
            with rc2:
                m_qty = st.number_input(f"Qty used (kg) #{i+1}", min_value=0.0, step=1.0, key=f"mat_qty_{i}")
            material_rows.append((m_label, m_qty))

        if st.button("➕ Add material row"):
            st.session_state["prod_material_rows"] += 1
            st.rerun()

        if st.button("✅ Save Production Record", use_container_width=True, type="primary"):
            if not prod_options or sel_product_label == "None":
                st.error("No active product selected.")
            elif qty_produced_in <= 0:
                st.error("Quantity produced must be greater than zero.")
            else:
                product_id = prod_options[sel_product_label]
                supervisor_id = sup_options.get(sel_supervisor)
                manager_id = mgr_options.get(sel_manager)
                production_id = run_execute(
                    """INSERT INTO production (production_date, product_id, qty_produced, qty_rejected,
                       supervisor_id, manager_id, labour_count, notes) VALUES (?,?,?,?,?,?,?,?)""",
                    (prod_date_in.strftime("%Y-%m-%d"), product_id, qty_produced_in, qty_rejected_in,
                     supervisor_id, manager_id, labour_count_in, prod_notes_in))

                for m_label, m_qty in material_rows:
                    if m_qty > 0 and m_label in mat_options:
                        mat_id = mat_options[m_label]
                        run_execute("INSERT INTO production_material_usage (production_id, material_id, qty_used) VALUES (?,?,?)",
                                    (production_id, mat_id, m_qty))
                        run_execute("UPDATE raw_materials SET stock_qty = stock_qty - ? WHERE material_id=?", (m_qty, mat_id))

                run_execute("UPDATE products SET stock_qty = stock_qty + ? WHERE product_id=?", (qty_produced_in, product_id))
                st.session_state["prod_material_rows"] = 1
                st.success(f"Production record #{production_id} saved successfully.")
                st.rerun()

    # ---------------------------- SCRAP / CONDEMNED MATERIAL ----------------------------
    elif page == "🗑️ Scrap / Condemned Material":
        st.subheader("🗑️ Scrap & Condemned Material Log")
        with st.form("scrap_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                scrap_date_in = st.date_input("Date", value=date.today())
                material_category_in = st.selectbox("Material Category", ["Copper", "PVC", "XLPE", "Aluminium"])
            with c2:
                sub_category_in = st.text_input("Sub-Category (e.g. Pure Copper, China Copper)")
                scrap_qty_in = st.number_input("Quantity (kg)", min_value=0.0, step=1.0)
            with c3:
                reason_in = st.selectbox("Reason", ["Condemned", "Handling Loss", "Broken in Production", "Other"])
                notes_in = st.text_input("Notes")
            log_scrap_btn = st.form_submit_button("Log Scrap Entry", use_container_width=True)

        if log_scrap_btn:
            if scrap_qty_in <= 0:
                st.error("Quantity must be greater than zero.")
            else:
                run_execute("""INSERT INTO scrap_log (log_date, material_category, sub_category, qty, unit, reason, notes)
                               VALUES (?,?,?,?,?,?,?)""",
                            (scrap_date_in.strftime("%Y-%m-%d"), material_category_in, sub_category_in, scrap_qty_in, "kg", reason_in, notes_in))
                st.success("Scrap entry logged.")
                st.rerun()

    # ---------------------------- ORDERS & SALES ----------------------------
    elif page == "🧾 Orders & Sales":
        st.subheader("🧾 Orders")
        orders_df = df_query("""SELECT o.order_id, c.name AS customer, c.phone, o.order_date, o.status, o.total_amount, o.notes
                                 FROM orders o JOIN customers c ON o.customer_id = c.customer_id
                                 ORDER BY o.order_id DESC""")
        st.dataframe(orders_df, use_container_width=True, hide_index=True)

    # ---------------------------- CUSTOMERS ----------------------------
    elif page == "👥 Customers":
        st.subheader("👥 Customer Records")
        cust_df = df_query("SELECT * FROM customers ORDER BY created_on DESC")
        st.dataframe(cust_df, use_container_width=True, hide_index=True)

    # ---------------------------- STAFF ----------------------------
    elif page == "🧑‍🏭 Staff":
        st.subheader("🧑‍🏭 Staff Records")
        staff_df = df_query("SELECT * FROM staff ORDER BY role, name")
        st.dataframe(staff_df, use_container_width=True, hide_index=True)

    # ---------------------------- REPORTS ----------------------------
    elif page == "📈 Reports":
        st.subheader("📈 Reports")
        products_df = df_query("SELECT category, material, spec_name, unit, price, stock_qty FROM products")
        st.download_button("⬇️ Download Product Stock Report (Excel)",
                            data=to_excel_bytes(products_df, "Product Stock"),
                            file_name=f"product_stock_{date.today()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
